import os
import glob
import re
import pandas as pd

def parse_report_file(filepath):
    """
    Parses a live volatile report .txt file and returns a dictionary of metrics.
    """
    with open(filepath, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    
    start_time_str = ''
    for line in lines:
        if 'Start Time' in line and ':' in line:
            start_time_str = line.split(':', 1)[1].strip()
            break
            
    if start_time_str:
        date_part = start_time_str.split()[0]
    else:
        m = re.search(r'(\d{8})_\d{6}', os.path.basename(filepath))
        if m:
            d = m.group(1)
            date_part = f"{d[:4]}-{d[4:6]}-{d[6:8]}"
        else:
            date_part = ''

    target_keys = [
        'Avg Return', 'Cumulative Return', 'Best Trade', 'Worst Trade',
        'Total P&L', 'CAGR (annualized)', 'Sharpe Ratio', 'Sortino Ratio',
        'Max Drawdown', 'Volatility (Ann.)', 'Win Rate', 'Win/Loss Ratio',
        'Avg Win', 'Avg Loss'
    ]

    market_condition = 'Neutral'
    for line in lines[:45]:
        if line.startswith('Market Condition') and ':' in line:
            market_condition = line.split(':', 1)[1].strip()
            break

    row_data = {
        'date': date_part,
        'Market Condition': market_condition,
        '_start_time': start_time_str
    }
    for key in target_keys:
        row_data[key] = ''

    section1_lines = lines[:45]
    for line in section1_lines:
        for key in target_keys:
            if line.startswith(key):
                rest = line[len(key):].strip()
                tokens = re.findall(r'[$₹]?[+-]?\d[\d,]*\.\d+%?|[$₹]?[+-]?\d+%?|--', rest)
                
                short_val = tokens[0] if len(tokens) >= 1 else ''
                long_val = tokens[1] if len(tokens) >= 2 else ''
                comb_val = tokens[2] if len(tokens) >= 3 else (tokens[-1] if tokens else '')
                
                val = comb_val
                if val == '--' or val == '':
                    if long_val != '--' and long_val != '':
                        val = long_val
                    elif short_val != '--' and short_val != '':
                        val = short_val
                
                row_data[key] = val
                break

    return row_data

def style_and_save_excel(df, save_paths):
    """
    Saves DataFrame to specified Excel file paths with clean formatting.
    """
    columns_order = [
        'date', 'Market Condition', 'Avg Return', 'Cumulative Return', 'Best Trade', 'Worst Trade',
        'Total P&L', 'CAGR (annualized)', 'Sharpe Ratio', 'Sortino Ratio',
        'Max Drawdown', 'Volatility (Ann.)', 'Win Rate', 'Win/Loss Ratio',
        'Avg Win', 'Avg Loss'
    ]
    
    if '_start_time' in df.columns:
        df = df.sort_values(by='_start_time', ascending=True)
        df = df.drop(columns=['_start_time'])
        
    df = df[columns_order]

    for path in save_paths:
        os.makedirs(os.path.dirname(os.path.abspath(path)), exist_ok=True)
        
        try:
            with pd.ExcelWriter(path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Volatile Results')
                workbook = writer.book
                worksheet = writer.sheets['Volatile Results']

                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
                cell_font = Font(name="Calibri", size=11)
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )
                
                for col_num, col_name in enumerate(columns_order, start=1):
                    cell = worksheet.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(columns_order)):
                    for cell_idx, cell in enumerate(row):
                        cell.font = cell_font
                        cell.border = thin_border
                        if cell_idx == 0:
                            cell.alignment = Alignment(horizontal="center")
                        else:
                            cell.alignment = Alignment(horizontal="right")

                for col in worksheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = col[0].column_letter
                    worksheet.column_dimensions[col_letter].width = max(max_len + 4, 12)

            print(f"Successfully saved Excel report to: {path}")
        except PermissionError:
            print(f"[WARNING] Could not save to {path} because the file is currently open in Excel or locked. Close the file to allow updating.")
        except Exception as e:
            print(f"[ERROR] Error saving Excel file {path}: {e}")

def update_volatile_excel(workspace_dir="."):
    """
    Main function to process US and Indian volatile report text files, export to Excel,
    and sync records to Google Sheets.
    """
    all_rows = []

    # 1. Process US Volatile Results
    us_dir = os.path.join(workspace_dir, "live_volatile_results")
    us_txt_files = sorted(glob.glob(os.path.join(us_dir, "US_VOLATILE_*.txt")))
    
    if us_txt_files:
        us_rows = [parse_report_file(f) for f in us_txt_files]
        for r in us_rows:
            r['Market'] = 'US'
        all_rows.extend(us_rows)
        df_us = pd.DataFrame(us_rows)
        us_save_paths = [
            os.path.join(us_dir, "live_volatiles_results_all_us.xlsx"),
            os.path.join(workspace_dir, "live_volatiles_results_all_us.xlsx")
        ]
        style_and_save_excel(df_us, us_save_paths)
    else:
        print(f"No US volatile report .txt files found in {us_dir}")

    # 2. Process Indian Volatile Results
    india_dir = os.path.join(workspace_dir, "Indian_log_volatile")
    india_txt_files = sorted(glob.glob(os.path.join(india_dir, "INDIA_VOLATILE_*.txt")))
    
    if india_txt_files:
        india_rows = [parse_report_file(f) for f in india_txt_files]
        for r in india_rows:
            r['Market'] = 'INDIA'
        all_rows.extend(india_rows)
        df_india = pd.DataFrame(india_rows)
        india_save_paths = [
            os.path.join(india_dir, "live_volatiles_results_all_india.xlsx"),
            os.path.join(workspace_dir, "live_volatiles_results_all_india.xlsx")
        ]
        style_and_save_excel(df_india, india_save_paths)
    else:
        print(f"No Indian volatile report .txt files found in {india_dir}")

    # 3. Sync to Google Sheets
    if all_rows:
        try:
            import sys
            root_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
            if root_dir not in sys.path:
                sys.path.insert(0, root_dir)
            from google_sheets_sync import sync_dataframe_to_tab
            df_all = pd.DataFrame(all_rows)
            sync_dataframe_to_tab("With LLM With Sentiment", df_all)
        except Exception as ex_gs:
            print(f"[WARNING] Could not sync to Google Sheets: {ex_gs}")

if __name__ == "__main__":
    update_volatile_excel()
